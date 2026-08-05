"""GET /event/metrics — the official dataset's own performance figures,
served verbatim so the UI never re-derives them (CLAUDE.md §2):

  hubs               Network_Performance, LATEST week, one entry per hub
  at_risk            hub ids whose official status is 'At Risk'
  baselines          the Baseline_Metrics table (metric / current / target / notes)
  weekly_demand      Σ weekly_volume per week — the 13-week demand line
  cost_per_shipment  the file's own fully-loaded AED/shipment per facility
                     (Cost_to_Serve: total_cost / shipments — THEIR arithmetic,
                     reproduced verbatim, labelled fully-loaded)
  network_volumes    latest-week Σ daily_avg per network type — so the UI's
                     all-networks headline is data, not a literal
"""

from pathlib import Path
from typing import Any

from fastapi import APIRouter, HTTPException

router = APIRouter()

RAW_WORKBOOK = Path(__file__).resolve().parents[2] / "data" / "dataset_g.xlsx"

HUB_FIELDS = [
    "courier_utilisation_pct",
    "vehicle_utilisation_pct",
    "on_time_delivery_pct",
    "first_attempt_success_pct",
    "capacity_headroom_pct",
    "sla_breach_count",
    "avg_delivery_time_min",
    "status",
]


@router.get("/event/metrics")
def get_event_metrics() -> dict[str, Any]:
    if not RAW_WORKBOOK.exists():
        raise HTTPException(404, "Event workbook not present (hubris/data/dataset_g.xlsx)")

    import openpyxl  # local: keep module import light

    wb = openpyxl.load_workbook(RAW_WORKBOOK, read_only=True, data_only=True)

    # ── Network_Performance: latest week, hub rows only ──
    perf = list(wb["Network_Performance"].iter_rows(min_row=2, values_only=True))
    latest = max(int(r[2]) for r in perf if r[2] is not None)
    hubs: dict[str, dict[str, Any]] = {}
    for r in perf:
        if r[2] is None or int(r[2]) != latest:
            continue
        facility = str(r[1])
        if not facility.startswith("HUB_"):
            continue
        hubs[facility] = dict(zip(HUB_FIELDS, r[4:12]))
    at_risk = sorted(h for h, m in hubs.items() if m["status"] == "At Risk")

    # ── Baseline_Metrics: the labelled table under the sheet preamble ──
    baselines: list[dict[str, Any]] = []
    seen_header = False
    for r in wb["Baseline_Metrics"].iter_rows(values_only=True):
        if r[0] == "Metric":
            seen_header = True
            continue
        if not seen_header or r[0] in (None, "DONE WHEN"):
            if r[0] == "DONE WHEN":
                break
            continue
        baselines.append(
            {"metric": str(r[0]), "current": str(r[1]), "target": str(r[2]), "notes": str(r[3])}
        )

    # ── Cost_to_Serve: the file's own fully-loaded per-facility figure ──
    facility_cost: dict[str, dict[str, float]] = {}
    for r in wb["Cost_to_Serve"].iter_rows(min_row=2, values_only=True):
        if r[1] is None:
            continue
        entry = facility_cost.setdefault(str(r[1]), {"ships": 0.0, "total": 0.0})
        entry["ships"] += float(r[3] or 0)
        entry["total"] += float(r[10] or 0)
    cost_per_shipment = {
        fac: round(v["total"] / v["ships"], 2)
        for fac, v in facility_cost.items()
        if v["ships"] > 0
    }

    # ── weekly demand: Σ weekly_volume per week (all network types) ──
    totals: dict[int, float] = {}
    for r in wb["Demand_by_Zone"].iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        week = int(r[0])
        totals[week] = totals.get(week, 0.0) + float(r[7] or 0)
    weekly_demand = [
        {"week": week, "total_volume": round(volume, 2)} for week, volume in sorted(totals.items())
    ]

    # ── latest-week daily volume per network type ──
    demand_rows = list(wb["Demand_by_Zone"].iter_rows(min_row=2, values_only=True))
    latest_demand_week = max(int(r[0]) for r in demand_rows if r[0] is not None)
    network_volumes: dict[str, float] = {}
    for r in demand_rows:
        if r[0] is None or int(r[0]) != latest_demand_week:
            continue
        network = str(r[2])
        network_volumes[network] = round(network_volumes.get(network, 0.0) + float(r[8] or 0), 2)

    return {
        "week": latest,
        "hub_count": len(hubs),
        "hubs": hubs,
        "at_risk": at_risk,
        "at_risk_count": len(at_risk),
        "baselines": baselines,
        "weekly_demand": weekly_demand,
        "cost_per_shipment": cost_per_shipment,
        "network_volumes": network_volumes,
    }
