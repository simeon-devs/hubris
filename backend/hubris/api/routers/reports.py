"""Report & export endpoints — engine-generated, downloadable, EDITABLE files.

Everything here is assembled from the SAME engine calls the interactive API
uses (get_network, GetKpisTool, OptimiseNetworkTool, GenerateDecisionBriefTool)
— no new computation paths, per CLAUDE.md §2. The frontend only downloads.

Three files:
  /export/report.xlsx   executive workbook (Summary/Hubs/Zones/Flows/Recommendations)
  /export/network.xlsx  ROUND-TRIP dataset — canonical column names, editable in
                        Excel and re-uploadable through POST /ingest unchanged
  /export/report.md     the decision brief as markdown
"""

import io
from datetime import datetime, timezone

from fastapi import APIRouter, HTTPException, Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from hubris.agents.tools.generate_brief import GenerateDecisionBriefTool
from hubris.agents.tools.get_kpis import GetKpisTool
from hubris.agents.tools.optimise_network import OptimiseNetworkTool
from hubris.api.routers.network import get_network
from hubris.api.state import state
from hubris.core.contracts import NetworkModel

router = APIRouter()

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
BRAND_RED = "E8112D"
HEADER_FILL = PatternFill("solid", fgColor=BRAND_RED)
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _model_or_404(scenario_id: str | None) -> NetworkModel:
    try:
        return state.get_model(scenario_id)
    except KeyError as exc:
        raise HTTPException(404, f"Unknown scenario_id: {scenario_id}") from exc


def _scenario_label(scenario_id: str | None) -> str:
    if scenario_id is None:
        return "Baseline"
    return state.scenario_labels.get(scenario_id, scenario_id)


def _style_header_row(ws: Worksheet) -> None:
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.freeze_panes = "A2"


def _autosize(ws: Worksheet) -> None:
    for column_cells in ws.columns:
        width = max((len(str(c.value)) for c in column_cells if c.value is not None), default=0)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(width + 3, 42)


def _sheet(wb: Workbook, title: str, header: list[str], rows: list[list], styled: bool = True) -> Worksheet:
    ws = wb.create_sheet(title)
    ws.append(header)
    for row in rows:
        ws.append(row)
    if styled:
        _style_header_row(ws)
    _autosize(ws)
    return ws


def _kpi_value(kpis: dict, key: str) -> float | None:
    value = kpis.get(key, {}).get("value")
    return value if isinstance(value, (int, float)) else None


KPI_KEYS = ["cost_to_serve", "utilization", "coverage", "spare_capacity"]


def _xlsx_response(wb: Workbook, filename: str) -> Response:
    buffer = io.BytesIO()
    wb.save(buffer)
    return Response(
        buffer.getvalue(),
        media_type=XLSX_MIME,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─────────────────────────────────────────────────────────────────────────────
# Executive report
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/export/report.xlsx")
def export_report(scenario_id: str | None = None, include_optimizer: bool = False) -> Response:
    model = _model_or_404(scenario_id)
    net = get_network(scenario_id)  # the exact payload the map renders
    kpis = GetKpisTool().run(model=model)

    wb = Workbook()
    wb.remove(wb.active)  # we name every sheet explicitly

    # ── Summary ──
    summary = wb.create_sheet("Summary")
    summary.append(["EMX ATLAS — Network Report"])
    summary["A1"].font = Font(bold=True, size=14, color=BRAND_RED)
    summary.append(["Generated at", datetime.now(timezone.utc).isoformat(timespec="seconds")])
    summary.append(["Scenario", _scenario_label(scenario_id)])
    summary.append(["Distance mode", net.distance_mode])
    summary.append([])
    if scenario_id is not None:
        base_kpis = GetKpisTool().run(model=state.baseline)
        summary.append(["KPI", "Baseline", "Scenario", "Δ vs baseline (%)", "Unit"])
        for cell in summary[summary.max_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for key in KPI_KEYS:
            base_v, scen_v = _kpi_value(base_kpis, key), _kpi_value(kpis, key)
            delta_pct = (
                round((scen_v - base_v) / base_v * 100, 2)
                if base_v not in (None, 0) and scen_v is not None
                else None
            )
            summary.append([key, base_v, scen_v, delta_pct, kpis.get(key, {}).get("unit", "")])
    else:
        summary.append(["KPI", "Value", "Unit"])
        for cell in summary[summary.max_row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for key in KPI_KEYS:
            summary.append([key, _kpi_value(kpis, key), kpis.get(key, {}).get("unit", "")])
    _autosize(summary)

    # ── Hubs / Zones / Flows — straight from the /network payload ──
    _sheet(
        wb,
        "Hubs",
        ["id", "name", "emirate", "lat", "lon", "capacity", "status", "utilization_pct",
         "spare_capacity", "cost_to_serve", "required_headcount", "gap"],
        [[h.id, h.name, h.emirate, h.lat, h.lon, h.capacity, h.status, h.utilization_pct,
          h.spare_capacity, h.cost_to_serve, h.required_headcount, h.headcount_gap]
         for h in net.hubs],
    )
    _sheet(
        wb,
        "Zones",
        ["id", "name", "emirate", "lat", "lon", "demand"],
        [[z.id, z.name, z.emirate, z.lat, z.lon, z.demand] for z in net.zones],
    )
    _sheet(
        wb,
        "Flows",
        ["hub_id", "zone_id", "volume"],
        [[f.hub_id, f.zone_id, f.volume] for f in net.flows],
    )

    # ── Recommendations ──
    if include_optimizer:
        rec = OptimiseNetworkTool().run(
            model=model, objective={}, constraints=[], optimizer_name="milp_cflp",
            demand_variation_pct=20.0,
        )
        rows = [[c.get("action", ""), c.get("hub_id", "")] for c in rec.get("changes", [])] or [
            ["already_optimal", ""]
        ]
        ws = _sheet(wb, "Recommendations", ["action", "hub_id"], rows)
        ws.append([])
        ws.append(["cost_to_serve_before", rec.get("cost_to_serve_before")])
        ws.append(["cost_to_serve_after", rec.get("cost_to_serve_after")])
        ws.append(["savings_per_parcel", rec.get("cost_to_serve_savings_per_parcel")])
        robustness = rec.get("robustness", {})
        ws.append(["robustness_p10_p90",
                   f"{robustness.get('cost_to_serve_p10')} – {robustness.get('cost_to_serve_p90')}"])
        ws.append(["feasible_pct", robustness.get("feasible_pct")])
        _autosize(ws)
    else:
        _sheet(
            wb,
            "Recommendations",
            ["note"],
            [["Optimizer not run — request with include_optimizer=true for the recommendation."]],
        )

    return _xlsx_response(wb, "atlas-report.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Round-trip network dataset
# ─────────────────────────────────────────────────────────────────────────────
@router.get("/export/network.xlsx")
def export_network(scenario_id: str | None = None) -> Response:
    """Canonical column names ONLY — this file must survive Excel editing and
    come back in through POST /ingest without any column_overrides."""
    model = _model_or_404(scenario_id)

    wb = Workbook()
    wb.remove(wb.active)
    _sheet(
        wb,
        "hubs",
        ["id", "name", "lat", "lon", "emirate", "capacity", "fixed_cost", "handling_cost", "status"],
        [[h.id, h.name, h.lat, h.lon, h.emirate, h.capacity, h.fixed_cost, h.handling_cost, h.status]
         for h in model.hubs],
        styled=False,
    )
    _sheet(
        wb,
        "zones",
        ["id", "name", "lat", "lon", "emirate", "demand", "sla_hours"],
        [[z.id, z.name, z.lat, z.lon, z.emirate, z.demand, z.sla_hours] for z in model.zones],
        styled=False,
    )
    _sheet(
        wb,
        "fleet_types",
        ["id", "name", "capacity", "cost_per_km", "fixed_cost", "count_available", "hub_id"],
        [[f.id, f.name, f.capacity, f.cost_per_km, f.fixed_cost, f.count_available, f.hub_id]
         for f in model.fleet_types],
        styled=False,
    )
    _sheet(
        wb,
        "od_matrix",
        ["from_id", "to_id", "distance_km", "time_min", "cost"],
        [[od.from_id, od.to_id, od.distance_km, od.time_min, od.cost]
         for od in model.od_matrix.values()],
        styled=False,
    )
    _sheet(
        wb,
        "current_assignments",
        ["zone_id", "hub_id", "volume"],
        [[zone_id, hub_id, model.demand.get(zone_id, 0.0)]
         for zone_id, hub_id in (model.assignments or {}).items()],
        styled=False,
    )
    return _xlsx_response(wb, "atlas-network.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# Decision brief as markdown
# ─────────────────────────────────────────────────────────────────────────────
def _md_value(value: object) -> str:
    if isinstance(value, float):
        return f"{value:,.4g}"
    return str(value)


def _md_section(title: str, payload: object) -> str:
    lines = [f"## {title}", ""]
    if payload is None:
        lines.append("_None._")
    elif isinstance(payload, dict):
        for key, value in payload.items():
            if isinstance(value, dict):
                lines.append(f"- **{key.replace('_', ' ')}**:")
                lines.extend(
                    f"  - {k.replace('_', ' ')}: {_md_value(v)}" for k, v in value.items()
                )
            elif isinstance(value, list):
                lines.append(f"- **{key.replace('_', ' ')}**: " + (
                    "; ".join(
                        " ".join(_md_value(v) for v in item.values()) if isinstance(item, dict) else _md_value(item)
                        for item in value
                    ) or "none"
                ))
            else:
                lines.append(f"- **{key.replace('_', ' ')}**: {_md_value(value)}")
    else:
        lines.append(_md_value(payload))
    lines.append("")
    return "\n".join(lines)


@router.get("/export/report.md")
def export_report_md(
    scenario_id: str | None = None,
    optimizer_name: str = "milp_cflp",
    demand_variation_pct: float = 20.0,
) -> Response:
    model = _model_or_404(scenario_id)
    brief = GenerateDecisionBriefTool().run(
        model=model, optimizer_name=optimizer_name, demand_variation_pct=demand_variation_pct
    )

    parts = [
        "# EMX ATLAS — Decision Brief",
        "",
        f"_Generated {brief.get('generated_at', '')} · Scenario: {_scenario_label(scenario_id)}_",
        "",
        "## Summary",
        "",
        str(brief.get("summary", "")),
        "",
        _md_section("Current state", brief.get("current_state")),
        _md_section("Proposed change", brief.get("proposed_change")),
        _md_section("Cost / Risk", brief.get("cost_risk")),
        _md_section("Sensitivity", brief.get("sensitivity")),
        _md_section("What it unblocks", brief.get("what_it_unblocks")),
    ]
    return Response(
        "\n".join(parts),
        media_type="text/markdown; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="atlas-brief.md"'},
    )
