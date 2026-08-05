"""Transform the official event workbook into the canonical schema.

Input : hubris/data/_event_raw.xlsx  (Dataset G, 11 sheets, as shipped)
Output: hubris/data/EMX_canonical.xlsx      — canonical 4-sheet workbook
        hubris/data/event_candidates.json   — Candidate hubs, real params

Every figure is copied or mechanically converted from dataset fields:
  hubs   ← Hub_Network (status=Active): capacity = max_daily_shipments,
           fixed_cost = monthly_rent_aed / 30 (per-day, engine unit),
           handling_cost = Σoverhead_cost_aed / Σmonthly_shipments from
           Cost_to_Serve for that hub (network median when absent).
  zones  ← Demand_by_Zone, LATEST week only, demand = Σ daily_avg across
           network types per (emirate, zone). sla_hours = 24.
           COORDINATES: the dataset ships none for zones — each zone is
           placed at its dominant serving facility plus a small
           deterministic offset (presentation geometry ONLY; demand values
           are verbatim).
  fleet_types ← Fleet_Roster grouped by vehicle_type (counts summed,
           per-km cost and capacity averaged, monthly cost / 30).
  current_assignments ← latest-week rows whose serving id is an Active hub.
  od_matrix: omitted on purpose — the ingestion engine derives it.

Run inside the backend container:
  python -m hubris.data.transform_event_dataset
"""

import json
import math
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
RAW = HERE / "_event_raw.xlsx"
CANONICAL = HERE / "EMX_canonical.xlsx"
CANDIDATES = HERE / "event_candidates.json"


def _rows(ws) -> list[dict]:
    lines = list(ws.iter_rows(values_only=True))
    headers = [str(h) for h in lines[0]]
    return [dict(zip(headers, line)) for line in lines[1:] if any(v is not None for v in line)]


def _offset(name: str, index: int) -> tuple[float, float]:
    """Deterministic small ring offset for zone geometry (no coords in data)."""
    angle = (hash(name) % 360) * math.pi / 180
    radius = 0.018 + 0.012 * (index % 3)
    return radius * math.cos(angle) * 1.25, radius * math.sin(angle) * 0.8


def main() -> None:
    wb = openpyxl.load_workbook(RAW, read_only=True, data_only=True)
    hub_rows = _rows(wb["Hub_Network"])
    store_rows = _rows(wb["Dark_Store_Network"])
    demand_rows = _rows(wb["Demand_by_Zone"])
    fleet_rows = _rows(wb["Fleet_Roster"])
    cost_rows = _rows(wb["Cost_to_Serve"])

    # ── handling cost per hub: overhead / shipments (dataset fields) ──
    overhead: dict[str, list[float]] = {}
    for r in cost_rows:
        hub_id = str(r["hub_or_store_id"])
        if r["monthly_shipments"]:
            overhead.setdefault(hub_id, [0.0, 0.0])
            overhead[hub_id][0] += float(r["overhead_cost_aed"] or 0)
            overhead[hub_id][1] += float(r["monthly_shipments"] or 0)
    handling = {
        hub_id: round(total / ship, 2) for hub_id, (total, ship) in overhead.items() if ship
    }
    median_handling = sorted(handling.values())[len(handling) // 2] if handling else 1.0

    # ── hubs (Active) + candidates (Candidate) ──
    hubs, candidates = [], []
    for r in hub_rows:
        entry = {
            "id": str(r["hub_id"]),
            "name": str(r["hub_name"]),
            "lat": float(r["lat"]),
            "lon": float(r["lng"]),
            "emirate": str(r["emirate"]),
            "capacity": float(r["max_daily_shipments"]),
            "fixed_cost": round(float(r["monthly_rent_aed"]) / 30, 2),
            "handling_cost": handling.get(str(r["hub_id"]), median_handling),
            "status": "open",
        }
        if str(r["status"]).strip().lower() == "candidate":
            candidates.append({**entry, "status": "candidate", "zone": str(r["zone"])})
        else:
            hubs.append(entry)

    facility_pos = {h["id"]: (h["lat"], h["lon"]) for h in hubs}
    for r in store_rows:
        facility_pos[str(r["store_id"])] = (float(r["lat"]), float(r["lng"]))
    default_pos = (
        sum(p[0] for p in facility_pos.values()) / len(facility_pos),
        sum(p[1] for p in facility_pos.values()) / len(facility_pos),
    )

    # ── zones from the LATEST week ──
    latest = max(int(r["week_number"]) for r in demand_rows)
    week = [r for r in demand_rows if int(r["week_number"]) == latest]
    zone_demand: dict[tuple[str, str], float] = {}
    zone_server: dict[tuple[str, str], dict[str, float]] = {}
    for r in week:
        key = (str(r["emirate"]), str(r["zone"]))
        daily = float(r["daily_avg"] or 0)
        zone_demand[key] = zone_demand.get(key, 0.0) + daily
        zone_server.setdefault(key, {})
        server = str(r["serving_hub_or_store_id"])
        zone_server[key][server] = zone_server[key].get(server, 0.0) + daily

    zones, assignments = [], []
    for index, ((emirate, zone), demand) in enumerate(sorted(zone_demand.items())):
        dominant = max(zone_server[(emirate, zone)].items(), key=lambda kv: kv[1])[0]
        base = facility_pos.get(dominant, default_pos)
        dlon, dlat = _offset(zone, index)
        zone_id = f"Z_{emirate[:3].upper()}_{index:03d}"
        zones.append({
            "id": zone_id,
            "name": zone,
            "lat": round(base[0] + dlat, 6),
            "lon": round(base[1] + dlon, 6),
            "emirate": emirate,
            "demand": round(demand, 2),
            "sla_hours": 24,
        })
        for server, volume in zone_server[(emirate, zone)].items():
            if server in {h["id"] for h in hubs}:
                assignments.append({"zone_id": zone_id, "hub_id": server, "volume": round(volume, 2)})

    # ── fleet types grouped by vehicle_type ──
    grouped: dict[str, list[dict]] = {}
    for r in fleet_rows:
        grouped.setdefault(str(r["vehicle_type"]), []).append(r)
    fleet_types = []
    for i, (vtype, rows) in enumerate(sorted(grouped.items()), start=1):
        count = sum(int(r["vehicle_count"] or 0) for r in rows)
        fleet_types.append({
            "id": f"F{i}",
            "name": vtype,
            "capacity": round(sum(float(r["avg_capacity_units"] or 0) for r in rows) / len(rows), 2),
            "cost_per_km": round(sum(float(r["fuel_cost_per_km_aed"] or 0) for r in rows) / len(rows), 3),
            "fixed_cost": round(
                sum(float(r["monthly_fleet_cost_per_vehicle_aed"] or 0) for r in rows) / len(rows) / 30, 2
            ),
            "count_available": count,
        })

    # ── write canonical workbook ──
    out = openpyxl.Workbook()
    out.remove(out.active)

    def sheet(title: str, header: list[str], rows: list[dict]) -> None:
        ws = out.create_sheet(title)
        ws.append(header)
        for row in rows:
            ws.append([row[h] for h in header])

    sheet("hubs", ["id", "name", "lat", "lon", "emirate", "capacity", "fixed_cost", "handling_cost", "status"], hubs)
    sheet("zones", ["id", "name", "lat", "lon", "emirate", "demand", "sla_hours"], zones)
    sheet("fleet_types", ["id", "name", "capacity", "cost_per_km", "fixed_cost", "count_available"],
          [{**f, "hub_id": None} for f in fleet_types])
    sheet("current_assignments", ["zone_id", "hub_id", "volume"], assignments)
    out.save(CANONICAL)

    CANDIDATES.write_text(json.dumps(candidates, indent=2))
    print(f"hubs={len(hubs)} candidates={len(candidates)} zones={len(zones)} "
          f"assignments={len(assignments)} fleet_types={len(fleet_types)} week={latest}")
    print(f"wrote {CANONICAL.name} + {CANDIDATES.name}")


if __name__ == "__main__":
    main()
