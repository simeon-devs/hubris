"""Tests for the report/export endpoints (reports router).

The crown jewel is the ROUND-TRIP: /export/network.xlsx must emit the
canonical column names the ingestion schema mapper expects, so a planner can
edit the file in Excel and re-upload it through POST /ingest unchanged.
"""

import io

import openpyxl
import pytest
from fastapi.testclient import TestClient

from hubris.agents.builder import builder
from hubris.api.main import app
from hubris.api.state import state as app_state

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@pytest.fixture(scope="module")
def client():
    with TestClient(app) as c:
        yield c


@pytest.fixture(autouse=True)
def _isolate_state():
    original_baseline = app_state.baseline
    original_scenarios = dict(app_state.scenarios)
    original_labels = dict(app_state.scenario_labels)
    original_distance_mode = app_state.distance_mode
    original_agents = dict(builder._agents)
    yield
    app_state.baseline = original_baseline
    app_state.scenarios = original_scenarios
    app_state.scenario_labels = original_labels
    app_state.distance_mode = original_distance_mode
    builder._agents = original_agents


def test_network_export_round_trips_through_ingest(client):
    """Export the baseline, re-upload the very same bytes, and the ingested
    counts must match — proof the file is edit-and-reupload ready."""
    before = client.get("/network").json()

    exported = client.get("/export/network.xlsx")
    assert exported.status_code == 200
    assert exported.headers["content-type"].startswith(XLSX_MIME)

    ingested = client.post(
        "/ingest", files={"file": ("network.xlsx", exported.content, XLSX_MIME)}
    )
    assert ingested.status_code == 200, ingested.text
    counts = ingested.json()
    assert counts["hubs"] == len(before["hubs"])
    assert counts["zones"] == len(before["zones"])
    assert counts["fleet_types"] == len(before["fleet_types"])
    # One dominant-hub assignment per zone (NetworkModel contract).
    assert counts["current_assignments"] == len(before["zones"])


def test_report_workbook_has_all_sheets_and_styling(client):
    response = client.get("/export/report.xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(XLSX_MIME)
    assert "attachment" in response.headers.get("content-disposition", "")

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    assert workbook.sheetnames == ["Summary", "Hubs", "Zones", "Flows", "Recommendations"]

    hubs = workbook["Hubs"]
    baseline_hub_count = len(client.get("/network").json()["hubs"])
    assert hubs.max_row == baseline_hub_count + 1  # + header
    # Styled: bold white header on 7X red, frozen top row.
    assert hubs.freeze_panes == "A2"
    assert hubs.cell(row=1, column=1).font.bold
    # Workforce columns made it in.
    header = [c.value for c in hubs[1]]
    assert "required_headcount" in header
    assert "gap" in header

    # Without include_optimizer the Recommendations sheet carries a note row.
    recs = workbook["Recommendations"]
    assert recs.max_row >= 1


def test_report_summary_includes_scenario_deltas(client):
    sim = client.post(
        "/simulate",
        json={
            "scenario_name": "demand_scale",
            "params": {"factor": 1.2},
            "save_as": "report-test",
        },
    )
    assert sim.status_code == 200

    response = client.get("/export/report.xlsx", params={"scenario_id": "report-test"})
    assert response.status_code == 200
    summary = openpyxl.load_workbook(io.BytesIO(response.content))["Summary"]
    cells = [str(c.value) for row in summary.iter_rows() for c in row if c.value is not None]
    assert any("report-test" in v for v in cells)  # scenario label present
    assert any("baseline" in v.lower() for v in cells)  # baseline column present


def test_report_md_is_a_markdown_attachment(client):
    response = client.get("/export/report.md")
    assert response.status_code == 200
    assert "attachment" in response.headers.get("content-disposition", "")
    body = response.text
    assert body.startswith("# ")
    assert "## Summary" in body


def test_export_unknown_scenario_404s(client):
    for path in ("/export/report.xlsx", "/export/network.xlsx", "/export/report.md"):
        assert client.get(path, params={"scenario_id": "nope"}).status_code == 404
